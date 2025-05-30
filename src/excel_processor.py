"""
Excel İşleme Modülü
Excel dosyalarını okuma ve analiz etme için optimize edilmiş
"""

import logging
import pandas as pd
import numpy as np
from typing import List, Dict, Optional, Any, Tuple
from dataclasses import dataclass
import openpyxl
from openpyxl import load_workbook
import xlrd
import os
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed

logger = logging.getLogger(__name__)

@dataclass
class SheetData:
    """Excel sayfa verisi container'ı"""
    sheet_name: str
    text_content: str
    raw_data: pd.DataFrame
    summary_stats: Dict
    metadata: Dict

@dataclass
class ExcelResult:
    """Excel işleme sonucu"""
    sheets: List[SheetData]
    total_sheets: int
    metadata: Dict

class ExcelProcessor:
    """
    Excel dosyaları için optimize edilmiş işlemci
    XLS, XLSX, XLSM formatlarını destekler
    Multiprocessing desteği ile sheet paralel işleme
    """
    
    def __init__(self, max_workers: int = None, use_multiprocessing: bool = True):
        """
        Excel işlemci başlatıcı
        
        Args:
            max_workers: Paralel işlem için maksimum worker sayısı
            use_multiprocessing: Multiprocessing kullanılıp kullanılmayacağı
        """
        self.max_workers = max_workers or min(4, os.cpu_count())
        self.use_multiprocessing = use_multiprocessing
        logger.info(f"📊 ExcelProcessor başlatıldı - {self.max_workers} worker, MP: {use_multiprocessing}")
    
    def process_excel(self, excel_path: str) -> ExcelResult:
        """
        Excel dosyasını işle ve sonuçları döndür
        
        Args:
            excel_path: Excel dosyasının yolu
            
        Returns:
            ExcelResult: İşlenmiş Excel verisi
        """
        logger.info(f"📊 Excel işleniyor: {excel_path}")
        
        try:
            # Dosya formatını belirle
            file_extension = os.path.splitext(excel_path)[1].lower()
            
            # Excel metadata'sını al
            excel_metadata = self._get_excel_metadata(excel_path, file_extension)
            
            # Sayfaları işle (paralel veya sıralı)
            sheets = self._process_all_sheets(excel_path, file_extension)
            
            result = ExcelResult(
                sheets=sheets,
                total_sheets=len(sheets),
                metadata=excel_metadata
            )
            
            logger.info(f"✅ Excel işlendi - {len(sheets)} sayfa")
            return result
            
        except Exception as e:
            logger.error(f"❌ Excel işleme hatası: {e}")
            raise
    
    def _get_excel_metadata(self, excel_path: str, file_extension: str) -> Dict:
        """Excel metadata'sını çıkar"""
        try:
            metadata = {
                'file_format': file_extension,
                'file_size': os.path.getsize(excel_path),
            }
            
            if file_extension in ['.xlsx', '.xlsm']:
                # openpyxl ile metadata al
                try:
                    wb = load_workbook(excel_path, read_only=True)
                    metadata.update({
                        'sheet_names': wb.sheetnames,
                        'sheet_count': len(wb.sheetnames),
                        'created': str(wb.properties.created) if wb.properties.created else '',
                        'modified': str(wb.properties.modified) if wb.properties.modified else '',
                        'creator': wb.properties.creator or '',
                        'title': wb.properties.title or ''
                    })
                    wb.close()
                except Exception as e:
                    logger.warning(f"⚠️ openpyxl metadata alınamadı: {e}")
            
            elif file_extension == '.xls':
                # xlrd ile metadata al
                try:
                    with xlrd.open_workbook(excel_path) as wb:
                        metadata.update({
                            'sheet_names': wb.sheet_names(),
                            'sheet_count': wb.nsheets
                        })
                except Exception as e:
                    logger.warning(f"⚠️ xlrd metadata alınamadı: {e}")
            
            return metadata
            
        except Exception as e:
            logger.warning(f"⚠️ Excel metadata alınamadı: {e}")
            return {'file_format': file_extension}
    
    def _process_all_sheets(self, excel_path: str, file_extension: str) -> List[SheetData]:
        """Tüm sayfaları işle (paralel veya sıralı)"""
        sheets = []
        
        try:
            # Pandas ile tüm sayfaları oku
            if file_extension == '.xls':
                # xlrd engine kullan
                all_sheets = pd.read_excel(excel_path, sheet_name=None, engine='xlrd')
            else:
                # openpyxl engine kullan
                all_sheets = pd.read_excel(excel_path, sheet_name=None, engine='openpyxl')
            
            sheet_count = len(all_sheets)
            logger.info(f"📊 {sheet_count} sayfa bulundu")
            
            # Paralel işleme kararı
            if sheet_count <= 2 or not self.use_multiprocessing:
                # Küçük Excel dosyaları için sıralı işleme
                logger.info("📄 Küçük Excel - sıralı işleme")
                for sheet_name, df in all_sheets.items():
                    logger.debug(f"📄 İşleniyor: {sheet_name}")
                    sheet_data = self._process_single_sheet(sheet_name, df)
                    if sheet_data:
                        sheets.append(sheet_data)
            else:
                # Büyük Excel dosyaları için paralel işleme
                logger.info(f"🚀 Paralel işleme başlatılıyor - {self.max_workers} worker")
                sheets = self._process_sheets_parallel(all_sheets)
            
            logger.info(f"✅ {len(sheets)} sayfa başarıyla işlendi")
            
        except Exception as e:
            logger.error(f"❌ Sayfalar işlenemedi: {e}")
            raise
        
        return sheets
    
    def _process_sheets_parallel(self, all_sheets: Dict[str, pd.DataFrame]) -> List[SheetData]:
        """Sayfaları paralel olarak işle"""
        sheets = []
        
        # ThreadPoolExecutor kullan (I/O bound operations için daha uygun)
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Tüm sayfalar için task'ları başlat
            future_to_sheet = {
                executor.submit(self._process_single_sheet, sheet_name, df): sheet_name
                for sheet_name, df in all_sheets.items()
            }
            
            # Sonuçları topla
            for future in as_completed(future_to_sheet):
                sheet_name = future_to_sheet[future]
                try:
                    sheet_data = future.result()
                    if sheet_data:
                        sheets.append(sheet_data)
                        logger.debug(f"✅ Sayfa işlendi: {sheet_name}")
                except Exception as e:
                    logger.error(f"❌ Sayfa işlenemedi {sheet_name}: {e}")
        
        # Sheet ismine göre sırala (orijinal sırayı koru)
        original_order = list(all_sheets.keys())
        sheets.sort(key=lambda x: original_order.index(x.sheet_name) if x.sheet_name in original_order else 999)
        
        return sheets
    
    def _process_single_sheet(self, sheet_name: str, df: pd.DataFrame) -> Optional[SheetData]:
        """Tek bir sayfayı işle"""
        try:
            # Boş sayfa kontrolü
            if df.empty:
                logger.debug(f"⚠️ Boş sayfa atlanıyor: {sheet_name}")
                return None
            
            # Veriyi temizle
            cleaned_df = self._clean_dataframe(df)
            
            # Çok az veri varsa atla
            if cleaned_df.empty or (cleaned_df.shape[0] <= 1 and cleaned_df.shape[1] <= 1):
                logger.debug(f"⚠️ Yetersiz veri, sayfa atlanıyor: {sheet_name}")
                return None
            
            # Özet istatistikleri hesapla
            summary_stats = self._calculate_summary_stats(cleaned_df)
            
            # Metin formatına çevir
            text_content = self._dataframe_to_text(cleaned_df, sheet_name)
            
            # Metadata oluştur
            metadata = {
                'original_shape': df.shape,
                'cleaned_shape': cleaned_df.shape,
                'data_types': dict(cleaned_df.dtypes.astype(str)),
                'has_numeric_data': any(cleaned_df.select_dtypes(include=[np.number]).columns),
                'has_text_data': any(cleaned_df.select_dtypes(include=['object']).columns),
                'null_counts': dict(cleaned_df.isnull().sum()),
                'processing_method': 'pandas'
            }
            
            return SheetData(
                sheet_name=sheet_name,
                text_content=text_content,
                raw_data=cleaned_df,
                summary_stats=summary_stats,
                metadata=metadata
            )
            
        except Exception as e:
            logger.error(f"❌ Sayfa işlenemedi {sheet_name}: {e}")
            return None
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """DataFrame'i temizle"""
        # Kopyasını al
        cleaned_df = df.copy()
        
        # Tamamen boş satır ve sütunları kaldır
        cleaned_df = cleaned_df.dropna(how='all').dropna(axis=1, how='all')
        
        # Index'i resetle
        cleaned_df = cleaned_df.reset_index(drop=True)
        
        # Sütun adlarını temizle
        if not cleaned_df.empty:
            # İlk satırı header olarak kullanmayı dene
            if cleaned_df.iloc[0].notna().any():
                new_columns = []
                for i, col in enumerate(cleaned_df.columns):
                    if pd.isna(cleaned_df.iloc[0, i]):
                        new_columns.append(f"Sütun_{i+1}")
                    else:
                        new_columns.append(str(cleaned_df.iloc[0, i]))
                
                # Eğer yeni column adları mantıklıysa kullan
                if len(set(new_columns)) > len(new_columns) * 0.7:  # %70 unique olmalı
                    cleaned_df.columns = new_columns
                    cleaned_df = cleaned_df.drop(cleaned_df.index[0]).reset_index(drop=True)
        
        return cleaned_df
    
    def _calculate_summary_stats(self, df: pd.DataFrame) -> Dict:
        """Özet istatistikleri hesapla"""
        stats = {
            'total_rows': len(df),
            'total_columns': len(df.columns),
            'numeric_columns': [],
            'text_columns': [],
            'null_percentage': 0
        }
        
        try:
            # Sayısal sütunlar
            numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
            stats['numeric_columns'] = numeric_cols
            stats['total_numeric_columns'] = len(numeric_cols)
            
            # Metin sütunları
            text_cols = df.select_dtypes(include=['object']).columns.tolist()
            stats['text_columns'] = text_cols
            stats['total_text_columns'] = len(text_cols)
            
            # Boş hücre yüzdesi
            total_cells = df.size
            null_cells = df.isnull().sum().sum()
            stats['null_percentage'] = (null_cells / total_cells * 100) if total_cells > 0 else 0
            
            # Sayısal sütunlar için özet
            if numeric_cols:
                numeric_summary = {}
                for col in numeric_cols:
                    try:
                        col_data = df[col].dropna()
                        if not col_data.empty:
                            numeric_summary[col] = {
                                'min': float(col_data.min()),
                                'max': float(col_data.max()),
                                'mean': float(col_data.mean()),
                                'sum': float(col_data.sum()),
                                'count': int(col_data.count())
                            }
                    except:
                        continue
                
                stats['numeric_summary'] = numeric_summary
            
        except Exception as e:
            logger.warning(f"⚠️ İstatistik hesaplanamadı: {e}")
        
        return stats
    
    def _dataframe_to_text(self, df: pd.DataFrame, sheet_name: str) -> str:
        """DataFrame'i aranabilir metin formatına çevir"""
        try:
            text_content = f"EXCEL SAYFASI: {sheet_name}\n\n"
            
            # Temel bilgiler
            text_content += f"Satır sayısı: {len(df)}\n"
            text_content += f"Sütun sayısı: {len(df.columns)}\n"
            text_content += f"Sütun adları: {', '.join(map(str, df.columns))}\n\n"
            
            # Veri örnekleri (ilk 10 satır)
            text_content += "VERİ İÇERİĞİ:\n"
            sample_data = df.head(10)
            
            # Sütun bazında veri
            for col in df.columns:
                text_content += f"\n{col}:\n"
                col_data = sample_data[col].dropna()
                if not col_data.empty:
                    values = col_data.astype(str).tolist()
                    text_content += ", ".join(values[:20])  # İlk 20 değer
                    if len(col_data) > 20:
                        text_content += f" ... (toplam {len(df[col].dropna())} değer)"
                text_content += "\n"
            
            # Sayısal özet
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                text_content += "\nSAYISAL ÖZET:\n"
                for col in numeric_cols:
                    col_data = df[col].dropna()
                    if not col_data.empty:
                        text_content += f"{col}: Min={col_data.min():.2f}, Max={col_data.max():.2f}, "
                        text_content += f"Ortalama={col_data.mean():.2f}, Toplam={col_data.sum():.2f}\n"
            
            # Tablo formatı (küçük tablolar için)
            if len(df) <= 20 and len(df.columns) <= 10:
                text_content += "\nTABLO FORMATI:\n"
                text_content += df.to_string(index=False, max_rows=20, max_cols=10)
                text_content += "\n"
            
            return text_content
            
        except Exception as e:
            logger.warning(f"⚠️ Metin dönüştürme hatası: {e}")
            return f"EXCEL SAYFASI: {sheet_name}\n[Metin dönüştürme hatası]"
    
    def get_excel_summary(self, excel_result: ExcelResult) -> Dict:
        """Excel dosyasının genel özetini çıkar"""
        summary = {
            'total_sheets': excel_result.total_sheets,
            'total_rows': 0,
            'total_columns': 0,
            'total_numeric_columns': 0,
            'sheet_summaries': []
        }
        
        try:
            for sheet in excel_result.sheets:
                sheet_summary = {
                    'name': sheet.sheet_name,
                    'rows': sheet.summary_stats.get('total_rows', 0),
                    'columns': sheet.summary_stats.get('total_columns', 0),
                    'numeric_columns': sheet.summary_stats.get('total_numeric_columns', 0),
                    'has_data': len(sheet.text_content) > 100
                }
                
                summary['sheet_summaries'].append(sheet_summary)
                summary['total_rows'] += sheet_summary['rows']
                summary['total_columns'] += sheet_summary['columns']
                summary['total_numeric_columns'] += sheet_summary['numeric_columns']
        
        except Exception as e:
            logger.warning(f"⚠️ Özet hesaplanamadı: {e}")
        
        return summary
    
    def get_processor_stats(self) -> Dict:
        """İşlemci istatistiklerini döndür"""
        return {
            'processor_type': 'ExcelProcessor',
            'max_workers': self.max_workers,
            'use_multiprocessing': self.use_multiprocessing,
            'supported_formats': ['.xls', '.xlsx', '.xlsm'],
            'features': [
                'multi_sheet_processing',
                'parallel_sheet_processing',
                'numeric_analysis',
                'text_extraction',
                'summary_statistics',
                'data_cleaning',
                'thread_based_concurrency'
            ],
            'engines': ['pandas', 'openpyxl', 'xlrd'],
            'optimization': {
                'small_files': 'sequential_processing (≤2 sheets)',
                'large_files': 'parallel_processing (>2 sheets)',
                'concurrency_type': 'ThreadPoolExecutor'
            }
        } 